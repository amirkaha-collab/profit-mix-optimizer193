# -*- coding: utf-8 -*-
"""
reporting/report_utils.py
─────────────────────────
Pure formatting and styling helpers for report generation.

No Streamlit. No app state. No file I/O.
All functions receive plain Python values and return strings or styled objects.
"""
from __future__ import annotations

import math
from datetime import date
from typing import Optional


# ── Number formatting ─────────────────────────────────────────────────────────

def fmt_pct(v, decimals: int = 1, fallback: str = "—") -> str:
    """Format a float as a percentage string."""
    try:
        f = float(v)
        if math.isnan(f):
            return fallback
        return f"{f:.{decimals}f}%"
    except (TypeError, ValueError):
        return fallback


def fmt_ils(v, fallback: str = "—") -> str:
    """Format a float as Israeli Shekel amount (e.g. ₪1,234,567)."""
    try:
        f = float(v)
        if math.isnan(f):
            return fallback
        return f"₪{f:,.0f}"
    except (TypeError, ValueError):
        return fallback


def fmt_float(v, decimals: int = 2, fallback: str = "—") -> str:
    try:
        f = float(v)
        if math.isnan(f):
            return fallback
        return f"{f:.{decimals}f}"
    except (TypeError, ValueError):
        return fallback


def fmt_delta(v, decimals: int = 1, suffix: str = "pp") -> str:
    """Format a signed delta value (e.g. +2.3pp)."""
    try:
        f = float(v)
        if math.isnan(f):
            return "—"
        sign = "+" if f >= 0 else ""
        return f"{sign}{f:.{decimals}f}{suffix}"
    except (TypeError, ValueError):
        return "—"


# ── Date helpers ──────────────────────────────────────────────────────────────

def today_str() -> str:
    return date.today().strftime("%Y-%m-%d")


def report_title_date(report_date: str) -> str:
    """Return a display-friendly date string for report headers."""
    if report_date:
        return report_date
    return date.today().strftime("%d/%m/%Y")


# ── Excel cell styling ────────────────────────────────────────────────────────

def _xl_header_fill():
    from openpyxl.styles import PatternFill
    return PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")


def _xl_header_font():
    from openpyxl.styles import Font
    return Font(color="FFFFFF", bold=True, size=11)


def _xl_subheader_fill():
    from openpyxl.styles import PatternFill
    return PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")


def _xl_subheader_font():
    from openpyxl.styles import Font
    return Font(color="1F3A5F", bold=True, size=10)


def _xl_alt_fill():
    from openpyxl.styles import PatternFill
    return PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")


def _xl_center_align():
    from openpyxl.styles import Alignment
    return Alignment(horizontal="center", vertical="center")


def _xl_right_align():
    from openpyxl.styles import Alignment
    return Alignment(horizontal="right", vertical="center")


def _xl_border_thin():
    from openpyxl.styles import Border, Side
    thin = Side(style="thin", color="D1D5DB")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_header_style(cell) -> None:
    cell.fill   = _xl_header_fill()
    cell.font   = _xl_header_font()
    cell.alignment = _xl_center_align()
    cell.border = _xl_border_thin()


def apply_subheader_style(cell) -> None:
    cell.fill      = _xl_subheader_fill()
    cell.font      = _xl_subheader_font()
    cell.alignment = _xl_right_align()
    cell.border    = _xl_border_thin()


def apply_data_style(cell, row_idx: int = 0) -> None:
    if row_idx % 2 == 1:
        cell.fill = _xl_alt_fill()
    cell.alignment = _xl_right_align()
    cell.border    = _xl_border_thin()


def apply_number_style(cell, row_idx: int = 0) -> None:
    if row_idx % 2 == 1:
        cell.fill = _xl_alt_fill()
    cell.alignment = _xl_center_align()
    cell.border    = _xl_border_thin()


# ── HTML report helpers ───────────────────────────────────────────────────────

_HTML_SKELETON = """\
<!DOCTYPE html>
<html dir="rtl" lang="he">
<head>
<meta charset="utf-8"/>
<title>{title}</title>
<style>
  body {{ font-family: 'Segoe UI', Arial, sans-serif; direction: rtl;
         background: #f5f7fa; color: #111827; margin: 0; padding: 24px; }}
  .report-header {{ background: #1F3A5F; color: #fff; border-radius: 10px;
                    padding: 20px 28px; margin-bottom: 24px; }}
  .report-header h1 {{ margin: 0; font-size: 22px; }}
  .report-header .sub {{ font-size: 13px; color: #93b4e0; margin-top: 4px; }}
  .section {{ background: #fff; border: 1px solid #e5eaf2; border-radius: 10px;
              padding: 18px 22px; margin-bottom: 18px; }}
  .section h2 {{ font-size: 15px; color: #1F3A5F; margin: 0 0 14px;
                 border-bottom: 2px solid #EFF6FF; padding-bottom: 8px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th {{ background: #1F3A5F; color: #fff; padding: 8px 12px; text-align: right; }}
  td {{ padding: 7px 12px; border-bottom: 1px solid #f1f5f9; text-align: right; }}
  tr:nth-child(even) td {{ background: #f9fafb; }}
  .kpi-row {{ display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 12px; }}
  .kpi {{ background: #EFF6FF; border: 1px solid #C7D7FF; border-radius: 8px;
           padding: 10px 16px; min-width: 110px; text-align: center; }}
  .kpi-val {{ font-size: 20px; font-weight: 800; color: #1F3A5F; }}
  .kpi-lbl {{ font-size: 11px; color: #6B7280; margin-top: 2px; }}
  .ai-block {{ background: #FFFBEB; border: 1px solid #FDE68A; border-radius: 8px;
               padding: 14px 18px; font-size: 13px; line-height: 1.6; }}
  .footer {{ font-size: 11px; color: #9CA3AF; text-align: center; margin-top: 30px; }}
</style>
</head>
<body>
{body}
<div class="footer">דוח זה הופק אוטומטית · {date}</div>
</body>
</html>
"""


def html_skeleton(title: str, body: str, date: str = "") -> str:
    return _HTML_SKELETON.format(
        title=title,
        body=body,
        date=date or today_str(),
    )


def html_kpi(label: str, value: str) -> str:
    return (
        f"<div class='kpi'>"
        f"<div class='kpi-val'>{value}</div>"
        f"<div class='kpi-lbl'>{label}</div>"
        f"</div>"
    )


def html_section(title: str, content: str) -> str:
    return f"<div class='section'><h2>{title}</h2>{content}</div>"


def html_table(headers: list[str], rows: list[list[str]]) -> str:
    ths = "".join(f"<th>{h}</th>" for h in headers)
    trs = ""
    for row in rows:
        tds = "".join(f"<td>{cell}</td>" for cell in row)
        trs += f"<tr>{tds}</tr>"
    return f"<table><thead><tr>{ths}</tr></thead><tbody>{trs}</tbody></table>"
