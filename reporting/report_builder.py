# -*- coding: utf-8 -*-
"""
reporting/report_builder.py
────────────────────────────
Report generation engine.

No Streamlit. No app state. No side effects beyond writing bytes.
Each function receives a typed input dataclass and returns bytes.

Public API
──────────
    build_portfolio_report(data: PortfolioReportInput)  -> bytes   (XLSX)
    build_portfolio_html(data: PortfolioReportInput)    -> str    (HTML)
    build_optimizer_report(data: OptimizerReportInput)  -> bytes   (XLSX)
    build_optimizer_html(data: OptimizerReportInput)    -> str    (HTML)
    build_isa_report(data: ISAReportInput)              -> bytes   (XLSX)
    build_isa_html(data: ISAReportInput)                -> str    (HTML)

All builders:
  - accept only the typed input dataclass
  - validate input before proceeding
  - return bytes (XLSX) or str (HTML)
  - raise ValueError with a clear Hebrew message on invalid input
"""
from __future__ import annotations

import io
import math
from typing import Optional

import numpy as np
import pandas as pd

from reporting.report_models import (
    ISAReportInput,
    OptimizerAlternative,
    OptimizerReportInput,
    PortfolioReportInput,
)
from reporting.report_utils import (
    fmt_delta,
    fmt_float,
    fmt_ils,
    fmt_pct,
    html_kpi,
    html_section,
    html_skeleton,
    html_table,
    report_title_date,
    apply_data_style,
    apply_header_style,
    apply_number_style,
    apply_subheader_style,
    today_str,
)


# ══════════════════════════════════════════════════════════════════════════════
# Portfolio / Client Report
# ══════════════════════════════════════════════════════════════════════════════

def build_portfolio_html(data: PortfolioReportInput) -> str:
    """Generate a self-contained HTML portfolio report."""
    warnings = data.validate()
    if warnings:
        raise ValueError("שגיאת קלט לדוח תיק: " + "; ".join(warnings))

    t = data.totals
    date_str = report_title_date(data.report_date)
    product_label = f" | {data.product_type}" if data.product_type else ""

    # ── KPI row ──────────────────────────────────────────────────────────────
    kpis = (
        html_kpi("שווי תיק", fmt_ils(t.get("total")))
        + html_kpi("מוצרים", str(t.get("n_products", "—")))
        + html_kpi("מנהלים", str(t.get("n_managers", "—")))
        + html_kpi('חו"ל', fmt_pct(t.get("foreign")))
        + html_kpi("מניות", fmt_pct(t.get("equity")))
        + html_kpi('מט"ח', fmt_pct(t.get("fx")))
        + html_kpi("לא סחיר", fmt_pct(t.get("illiquid")))
    )
    if not math.isnan(t.get("cost", math.nan)):
        kpis += html_kpi("עלות שנתית", fmt_pct(t.get("cost"), decimals=2))

    kpi_section = html_section("סיכום תיק", f"<div class='kpi-row'>{kpis}</div>")

    # ── Holdings table ────────────────────────────────────────────────────────
    df = data.holdings_df.copy()
    active = df[~df.get("excluded", pd.Series([False] * len(df))).astype(bool)]
    total_amt = float(active["amount"].sum()) if "amount" in active.columns else 0

    headers = ["גוף", "מוצר", "מסלול", "סוג", "סכום", "משקל %",
               "מניות %", 'חו"ל %', 'מט"ח %', "לא סחיר %"]
    rows = []
    for _, row in active.iterrows():
        amt = float(row.get("amount", 0) or 0)
        weight = round(amt / total_amt * 100, 1) if total_amt > 0 else 0
        rows.append([
            str(row.get("provider", "—")),
            str(row.get("product_name", "—")),
            str(row.get("track", "—")),
            str(row.get("product_type", "—")),
            fmt_ils(amt),
            f"{weight:.1f}%",
            fmt_pct(row.get("equity_pct")),
            fmt_pct(row.get("foreign_pct")),
            fmt_pct(row.get("fx_pct")),
            fmt_pct(row.get("illiquid_pct")),
        ])
    holdings_section = html_section("פירוט אחזקות", html_table(headers, rows))

    # ── AI commentary ─────────────────────────────────────────────────────────
    ai_section = ""
    if data.ai_commentary:
        ai_section = html_section(
            "ניתוח AI",
            f"<div class='ai-block'>{data.ai_commentary.replace(chr(10), '<br>')}</div>",
        )

    header = (
        f"<div class='report-header'>"
        f"<h1>📊 דוח תיק השקעות — {data.client_name}</h1>"
        f"<div class='sub'>{date_str}{product_label}</div>"
        f"</div>"
    )
    body = header + kpi_section + holdings_section + ai_section
    return html_skeleton(f"דוח תיק — {data.client_name}", body, date_str)


def build_portfolio_report(data: PortfolioReportInput) -> bytes:
    """Generate an XLSX portfolio report. Returns bytes."""
    warnings = data.validate()
    if warnings:
        raise ValueError("שגיאת קלט לדוח תיק: " + "; ".join(warnings))

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "תיק השקעות"
    ws.sheet_view.rightToLeft = True

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"דוח תיק השקעות — {data.client_name}  |  {report_title_date(data.report_date)}"
    apply_header_style(title_cell)
    ws.row_dimensions[1].height = 28

    if data.product_type:
        ws.merge_cells("A2:J2")
        sub = ws["A2"]
        sub.value = f"סוג מוצר: {data.product_type}"
        apply_subheader_style(sub)

    # ── KPI summary ────────────────────────────────────────────────────────────
    t = data.totals
    start_row = 3 if not data.product_type else 4
    kpi_labels = ["שווי תיק", "מוצרים", "מנהלים", 'חו"ל %', "מניות %", 'מט"ח %', "לא סחיר %"]
    kpi_values = [
        fmt_ils(t.get("total")),
        str(t.get("n_products", "—")),
        str(t.get("n_managers", "—")),
        fmt_pct(t.get("foreign")),
        fmt_pct(t.get("equity")),
        fmt_pct(t.get("fx")),
        fmt_pct(t.get("illiquid")),
    ]
    for i, (lbl, val) in enumerate(zip(kpi_labels, kpi_values), start=1):
        lbl_cell = ws.cell(row=start_row, column=i, value=lbl)
        apply_subheader_style(lbl_cell)
        val_cell = ws.cell(row=start_row + 1, column=i, value=val)
        apply_number_style(val_cell, 0)

    # ── Holdings table ─────────────────────────────────────────────────────────
    data_start = start_row + 3
    col_defs = [
        ("גוף",       "provider",      18),
        ("מוצר",      "product_name",  32),
        ("מסלול",     "track",         14),
        ("סוג",       "product_type",  16),
        ("סכום (₪)",  "amount",        14),
        ("משקל %",    None,            10),
        ("מניות %",   "equity_pct",    10),
        ('חו"ל %',    "foreign_pct",   10),
        ('מט"ח %',    "fx_pct",        10),
        ("לא סחיר %", "illiquid_pct",  12),
    ]
    headers = [c[0] for c in col_defs]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=data_start, column=ci, value=h)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(ci)].width = col_defs[ci - 1][2]

    df = data.holdings_df.copy()
    active = df[~df.get("excluded", pd.Series([False] * len(df))).astype(bool)]
    total_amt = float(active["amount"].sum()) if "amount" in active.columns else 0

    for ri, (_, row) in enumerate(active.iterrows()):
        xr = data_start + 1 + ri
        amt = float(row.get("amount", 0) or 0)
        weight = round(amt / total_amt * 100, 1) if total_amt > 0 else 0
        values = [
            str(row.get("provider", "")),
            str(row.get("product_name", "")),
            str(row.get("track", "")),
            str(row.get("product_type", "")),
            round(amt, 2),
            weight,
            _safe_float(row.get("equity_pct")),
            _safe_float(row.get("foreign_pct")),
            _safe_float(row.get("fx_pct")),
            _safe_float(row.get("illiquid_pct")),
        ]
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=xr, column=ci, value=val)
            if ci <= 4:
                apply_data_style(cell, ri)
            else:
                apply_number_style(cell, ri)
                if isinstance(val, float) and not math.isnan(val):
                    cell.number_format = '#,##0.00' if ci == 5 else '0.00'

    # ── AI commentary sheet ────────────────────────────────────────────────────
    if data.ai_commentary:
        ws_ai = wb.create_sheet("ניתוח AI")
        ws_ai.sheet_view.rightToLeft = True
        ws_ai.merge_cells("A1:D1")
        hdr = ws_ai["A1"]
        hdr.value = "ניתוח AI — סיכום"
        apply_header_style(hdr)
        for li, line in enumerate(data.ai_commentary.splitlines(), start=2):
            ws_ai.cell(row=li, column=1, value=line)
            ws_ai.column_dimensions["A"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# Optimizer Report
# ══════════════════════════════════════════════════════════════════════════════

def build_optimizer_html(data: OptimizerReportInput) -> str:
    """Generate a self-contained HTML optimizer results report."""
    warnings = data.validate()
    if warnings:
        raise ValueError("שגיאת קלט לדוח אופטימיזציה: " + "; ".join(warnings))

    date_str = report_title_date(data.report_date)
    product_label = f" | {data.product_type}" if data.product_type else ""

    # ── Target summary ────────────────────────────────────────────────────────
    target_kpis = ""
    for k, v in data.targets.items():
        try:
            if not math.isnan(float(v)):
                target_kpis += html_kpi(k, fmt_pct(v))
        except (TypeError, ValueError):
            target_kpis += html_kpi(k, str(v))
    target_section = html_section(
        "יעדי חשיפה",
        f"<div class='kpi-row'>{target_kpis}</div>",
    )

    # ── Alternatives table ────────────────────────────────────────────────────
    headers = ["חלופה", "מנהלים", "משקלים", 'חו"ל %', "מניות %", 'מט"ח %', "לא סחיר %", "שארפ", "שירות", "יתרון"]
    rows = []
    for alt in data.alternatives:
        weights_str = " / ".join(f"{w}%" for w in alt.weights) if alt.weights else "—"
        rows.append([
            alt.label,
            alt.managers.replace("|", " | "),
            weights_str,
            fmt_pct(alt.foreign_pct),
            fmt_pct(alt.stocks_pct),
            fmt_pct(alt.fx_pct),
            fmt_pct(alt.illiquid_pct),
            fmt_float(alt.sharpe),
            fmt_float(alt.service, decimals=1),
            alt.advantage,
        ])
    alts_section = html_section("חלופות מומלצות", html_table(headers, rows))

    # ── Baseline comparison (if provided) ────────────────────────────────────
    baseline_section = ""
    if data.baseline:
        bl = data.baseline
        bl_kpis = (
            html_kpi('חו"ל (קיים)', fmt_pct(bl.get("foreign")))
            + html_kpi("מניות (קיים)", fmt_pct(bl.get("stocks")))
            + html_kpi('מט"ח (קיים)', fmt_pct(bl.get("fx")))
            + html_kpi("לא סחיר (קיים)", fmt_pct(bl.get("illiquid")))
        )
        baseline_section = html_section(
            "השוואה למצב קיים",
            f"<div class='kpi-row'>{bl_kpis}</div>",
        )

    header = (
        f"<div class='report-header'>"
        f"<h1>⚡ דוח אופטימיזציה — {data.client_name}</h1>"
        f"<div class='sub'>{date_str}{product_label} | דירוג ראשי: {data.primary_rank}</div>"
        f"</div>"
    )
    body = header + target_section + alts_section + baseline_section
    return html_skeleton(f"אופטימיזציה — {data.client_name}", body, date_str)


def build_optimizer_report(data: OptimizerReportInput) -> bytes:
    """Generate an XLSX optimizer results report. Returns bytes."""
    warnings = data.validate()
    if warnings:
        raise ValueError("שגיאת קלט לדוח אופטימיזציה: " + "; ".join(warnings))

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "חלופות מומלצות"
    ws.sheet_view.rightToLeft = True

    # Title
    ws.merge_cells("A1:J1")
    tc = ws["A1"]
    tc.value = f"דוח אופטימיזציה — {data.client_name}  |  {report_title_date(data.report_date)}"
    apply_header_style(tc)
    ws.row_dimensions[1].height = 28

    # Targets row
    ws.merge_cells("A2:J2")
    tgt_str = "  ".join(f"{k}: {fmt_pct(v)}" for k, v in data.targets.items())
    tgt_cell = ws["A2"]
    tgt_cell.value = f"יעדים: {tgt_str}"
    apply_subheader_style(tgt_cell)

    # Alternatives headers
    col_defs = [
        ("חלופה",    12), ("מנהלים",  28), ("קופות",   30), ("מסלולים", 20),
        ("משקלים",   14), ('חו"ל %',  10), ("מניות %", 10), ('מט"ח %',  10),
        ("לא סחיר %", 12), ("שארפ",    9), ("שירות",   9),  ("יתרון",   24),
    ]
    for ci, (h, w) in enumerate(col_defs, start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, alt in enumerate(data.alternatives):
        xr = 4 + ri
        weights_str = " / ".join(f"{w}%" for w in alt.weights) if alt.weights else "—"
        values = [
            alt.label,
            alt.managers.replace("|", " | "),
            alt.funds.replace("|", " | "),
            alt.tracks.replace("|", " | "),
            weights_str,
            _safe_float(alt.foreign_pct),
            _safe_float(alt.stocks_pct),
            _safe_float(alt.fx_pct),
            _safe_float(alt.illiquid_pct),
            _safe_float(alt.sharpe),
            _safe_float(alt.service),
            alt.advantage,
        ]
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=xr, column=ci, value=val)
            if ci <= 5 or ci == 12:
                apply_data_style(cell, ri)
            else:
                apply_number_style(cell, ri)
                if isinstance(val, float) and not math.isnan(val):
                    cell.number_format = "0.00"

        # AI text for this alternative
        if alt.ai_text:
            ai_cell = ws.cell(row=xr, column=len(col_defs) + 1, value=alt.ai_text)
            apply_data_style(ai_cell, ri)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# ISA / Institutional Strategy Report
# ══════════════════════════════════════════════════════════════════════════════

def build_isa_html(data: ISAReportInput) -> str:
    """Generate a self-contained HTML ISA report."""
    warnings = data.validate()
    if warnings:
        raise ValueError("שגיאת קלט לדוח מוסדי: " + "; ".join(warnings))

    date_str = report_title_date(data.report_date)
    dr = data.date_range or _infer_date_range(data.display_df)

    # ── Metadata KPIs ──────────────────────────────────────────────────────────
    kpis = (
        html_kpi("גופים", str(len(data.managers)))
        + html_kpi("מסלולים", str(len(data.tracks)))
        + html_kpi("רכיבים", str(len(data.allocation_names)))
        + html_kpi("טווח", dr)
    )
    meta_section = html_section("פרמטרי הניתוח", f"<div class='kpi-row'>{kpis}</div>")

    # ── Latest snapshot table ─────────────────────────────────────────────────
    snap = _latest_snapshot(data.display_df)
    if not snap.empty:
        headers = ["גוף", "מסלול", "רכיב", "ערך (%)"]
        rows = [
            [
                str(row.get("manager", "")),
                str(row.get("track", "")),
                str(row.get("allocation_name", "")),
                fmt_pct(row.get("allocation_value"), decimals=2),
            ]
            for _, row in snap.iterrows()
        ]
        snap_section = html_section("ערכים עדכניים (Snapshot)", html_table(headers, rows))
    else:
        snap_section = ""

    # ── AI analysis sections ──────────────────────────────────────────────────
    ai_parts = ""
    for title, body in data.ai_sections.items():
        if body.strip():
            ai_parts += f"<h3 style='color:#1F3A5F;margin:16px 0 6px'>{title}</h3>"
            ai_parts += f"<div class='ai-block'>{body.replace(chr(10), '<br>')}</div>"
    ai_section = html_section("ניתוח AI", ai_parts) if ai_parts else ""

    product_label = f" | {data.product_type}" if data.product_type else ""
    header = (
        f"<div class='report-header'>"
        f"<h1>📐 דוח אסטרטגיות מוסדיים</h1>"
        f"<div class='sub'>{date_str}{product_label} | גופים: {', '.join(data.managers)}</div>"
        f"</div>"
    )
    body = header + meta_section + snap_section + ai_section
    return html_skeleton("דוח אסטרטגיות מוסדיים", body, date_str)


def build_isa_report(data: ISAReportInput) -> bytes:
    """Generate an XLSX ISA report. Returns bytes."""
    warnings = data.validate()
    if warnings:
        raise ValueError("שגיאת קלט לדוח מוסדי: " + "; ".join(warnings))

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Latest snapshot ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Snapshot"
    ws1.sheet_view.rightToLeft = True
    dr = data.date_range or _infer_date_range(data.display_df)

    ws1.merge_cells("A1:F1")
    tc = ws1["A1"]
    tc.value = f"דוח אסטרטגיות מוסדיים  |  {report_title_date(data.report_date)}  |  {dr}"
    apply_header_style(tc)
    ws1.row_dimensions[1].height = 28

    snap_headers = ["גוף", "מסלול", "רכיב", "ערך (%)", "תאריך"]
    col_widths   = [20,     14,       18,      10,        12]
    for ci, (h, w) in enumerate(zip(snap_headers, col_widths), start=1):
        cell = ws1.cell(row=2, column=ci, value=h)
        apply_header_style(cell)
        ws1.column_dimensions[get_column_letter(ci)].width = w

    snap = _latest_snapshot(data.display_df)
    for ri, (_, row) in enumerate(snap.iterrows()):
        xr = 3 + ri
        values = [
            str(row.get("manager", "")),
            str(row.get("track", "")),
            str(row.get("allocation_name", "")),
            _safe_float(row.get("allocation_value")),
            str(row["date"].strftime("%Y-%m") if pd.notna(row.get("date")) else ""),
        ]
        for ci, val in enumerate(values, start=1):
            cell = ws1.cell(row=xr, column=ci, value=val)
            if ci == 4:
                apply_number_style(cell, ri)
                if isinstance(val, float) and not math.isnan(val):
                    cell.number_format = "0.00"
            else:
                apply_data_style(cell, ri)

    # ── Sheet 2: Full time-series ──────────────────────────────────────────────
    ws2 = wb.create_sheet("סדרת זמן")
    ws2.sheet_view.rightToLeft = True
    ts_headers = ["גוף", "מסלול", "רכיב", "תאריך", "ערך (%)", "תדירות"]
    ts_widths   = [20,    14,      18,      12,       10,        10]
    for ci, (h, w) in enumerate(zip(ts_headers, ts_widths), start=1):
        cell = ws2.cell(row=1, column=ci, value=h)
        apply_header_style(cell)
        ws2.column_dimensions[get_column_letter(ci)].width = w

    df_sorted = data.display_df.sort_values(["manager", "track", "allocation_name", "date"])
    for ri, (_, row) in enumerate(df_sorted.iterrows()):
        xr = 2 + ri
        dt = row.get("date")
        dt_str = dt.strftime("%Y-%m") if pd.notna(dt) else ""
        values = [
            str(row.get("manager", "")),
            str(row.get("track", "")),
            str(row.get("allocation_name", "")),
            dt_str,
            _safe_float(row.get("allocation_value")),
            str(row.get("frequency", "")),
        ]
        for ci, val in enumerate(values, start=1):
            cell = ws2.cell(row=xr, column=ci, value=val)
            if ci == 5:
                apply_number_style(cell, ri)
                if isinstance(val, float) and not math.isnan(val):
                    cell.number_format = "0.00"
            else:
                apply_data_style(cell, ri)

    # ── Sheet 3: AI analysis sections ─────────────────────────────────────────
    if data.ai_sections:
        ws3 = wb.create_sheet("ניתוח AI")
        ws3.sheet_view.rightToLeft = True
        ws3.column_dimensions["A"].width = 20
        ws3.column_dimensions["B"].width = 80
        ws3.cell(row=1, column=1, value="כותרת").font = __import__("openpyxl").styles.Font(bold=True)
        ws3.cell(row=1, column=2, value="תוכן").font   = __import__("openpyxl").styles.Font(bold=True)
        ri = 2
        for title, body in data.ai_sections.items():
            if body.strip():
                ws3.cell(row=ri, column=1, value=title)
                ws3.cell(row=ri, column=2, value=body)
                ws3.row_dimensions[ri].height = max(15, body.count("\n") * 15 + 15)
                ri += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Private helpers ───────────────────────────────────────────────────────────

def _safe_float(v) -> float:
    """Convert to float, returning nan for missing/invalid values."""
    if v is None:
        return float("nan")
    try:
        f = float(v)
        return f
    except (TypeError, ValueError):
        return float("nan")


def _latest_snapshot(df: pd.DataFrame) -> pd.DataFrame:
    """Return the most recent value per manager/track/allocation_name."""
    if df.empty:
        return pd.DataFrame()
    try:
        grp_cols = ["manager", "track", "allocation_name"]
        idx = df.groupby(grp_cols)["date"].idxmax()
        snap = df.loc[idx].sort_values(grp_cols).reset_index(drop=True)
        return snap
    except Exception:
        return df.head(100)


def _infer_date_range(df: pd.DataFrame) -> str:
    try:
        return (
            f"{df['date'].min().strftime('%Y-%m')} – "
            f"{df['date'].max().strftime('%Y-%m')}"
        )
    except Exception:
        return "—"


# ══════════════════════════════════════════════════════════════════════════════
# Portfolio Comparison Builder
# ══════════════════════════════════════════════════════════════════════════════

def generate_html_report(
    comparison: "PortfolioComparison",
    actions: "list[PortfolioAction]",
) -> str:
    """
    Generate a self-contained Hebrew HTML portfolio proposal report.

    Parameters
    ──────────
    comparison : PortfolioComparison — current vs proposed snapshots with deltas
    actions    : list[PortfolioAction] — recommended changes (replace / add / remove)

    Returns
    ───────
    HTML string — fully self-contained, no external dependencies.

    Layout
    ──────
    1. Title header
    2. Executive Summary — KPI cards: allocations, sharpe, cost
    3. Allocation Table — before / after / delta for every allocation key
    4. Actions Section — one row per action
    """
    from reporting.report_models import PortfolioComparison, PortfolioAction
    from reporting.report_utils import fmt_pct, fmt_ils, fmt_delta, fmt_float, today_str

    cur  = comparison.current_snapshot
    prop = comparison.proposed_snapshot

    # ── Allocation labels (Hebrew display names) ──────────────────────────────
    _ALLOC_LABELS = {
        "equities": "מניות",
        "abroad":   'חו"ל',
        "fx":       'מט"ח',
        "illiquid": "לא סחיר",
    }
    all_keys = list({**cur.allocations, **prop.allocations}.keys())
    # Prefer canonical order; append anything extra
    _ORDER = ["equities", "abroad", "fx", "illiquid"]
    all_keys = [k for k in _ORDER if k in all_keys] + [k for k in all_keys if k not in _ORDER]

    def _delta_color(v: float) -> str:
        if math.isnan(v):     return "#6b7280"   # gray — unknown
        if v >  0.05:         return "#16a34a"   # green — improvement / increase
        if v < -0.05:         return "#dc2626"   # red   — decline / decrease
        return "#6b7280"                          # gray  — negligible

    def _delta_html(v: float, suffix: str = "pp") -> str:
        color = _delta_color(v)
        text  = fmt_delta(v, suffix=suffix)
        return f"<span style='color:{color};font-weight:700'>{text}</span>"

    # ── CSS ───────────────────────────────────────────────────────────────────
    css = """
<style>
*, *::before, *::after { box-sizing: border-box; }
body {
  font-family: 'Segoe UI', Arial, sans-serif;
  direction: rtl; background: #f5f7fa;
  color: #111827; margin: 0; padding: 28px;
}
.header {
  background: linear-gradient(135deg, #1F3A5F 0%, #2563eb 100%);
  color: #fff; border-radius: 12px; padding: 22px 28px; margin-bottom: 24px;
}
.header h1 { margin: 0; font-size: 24px; letter-spacing: -0.3px; }
.header .sub { font-size: 13px; color: #93c5fd; margin-top: 5px; }
.section {
  background: #fff; border: 1px solid #e5eaf2; border-radius: 10px;
  padding: 20px 24px; margin-bottom: 20px;
}
.section h2 {
  margin: 0 0 16px; font-size: 16px; color: #1F3A5F;
  border-bottom: 2px solid #EFF6FF; padding-bottom: 8px;
}
/* KPI cards */
.kpi-row { display: flex; gap: 12px; flex-wrap: wrap; }
.kpi {
  border-radius: 10px; padding: 12px 18px; min-width: 130px; text-align: center;
  border: 1px solid;
}
.kpi-cur  { background: #f3f4f6; border-color: #d1d5db; }
.kpi-prop { background: #f0fdf4; border-color: #86efac; }
.kpi-val  { font-size: 22px; font-weight: 900; line-height: 1.1; }
.kpi-cur  .kpi-val { color: #374151; }
.kpi-prop .kpi-val { color: #15803d; }
.kpi-lbl  { font-size: 11px; color: #6b7280; margin-top: 3px; }
.kpi-kind { font-size: 10px; font-weight: 600; margin-bottom: 2px; opacity: 0.7; }
/* Allocation table */
table { width: 100%; border-collapse: collapse; font-size: 13.5px; }
th {
  padding: 9px 14px; text-align: right; font-weight: 700; font-size: 12px;
  white-space: nowrap;
}
th.col-cur  { background: #f3f4f6; color: #374151; }
th.col-prop { background: #f0fdf4; color: #166534; }
th.col-delta { background: #fafafa; color: #6b7280; }
th.col-name  { background: #1F3A5F; color: #fff; }
td {
  padding: 9px 14px; border-bottom: 1px solid #f1f5f9; text-align: right;
  vertical-align: middle;
}
tr:last-child td { border-bottom: none; }
tr:hover td { background: #fafafa; }
td.name { font-weight: 700; color: #1F3A5F; }
/* Actions */
.action-row {
  display: flex; align-items: flex-start; gap: 12px;
  padding: 12px 0; border-bottom: 1px solid #f1f5f9;
}
.action-row:last-child { border-bottom: none; }
.badge {
  border-radius: 999px; padding: 3px 10px; font-size: 11px;
  font-weight: 700; white-space: nowrap; flex-shrink: 0; margin-top: 2px;
}
.badge-replace { background: #fef3c7; color: #92400e; }
.badge-add     { background: #dcfce7; color: #166534; }
.badge-remove  { background: #fee2e2; color: #991b1b; }
.action-text   { font-size: 13.5px; }
.action-impact { font-size: 12px; color: #6b7280; margin-top: 3px; }
.footer { font-size: 11px; color: #9ca3af; text-align: center; margin-top: 28px; }
</style>"""

    # ── 1. Header ─────────────────────────────────────────────────────────────
    header = f"""
<div class="header">
  <h1>📊 הצעת תיק השקעות</h1>
  <div class="sub">הופק: {today_str()}</div>
</div>"""

    # ── 2. Executive Summary ──────────────────────────────────────────────────
    def _kpi_pair(label: str, cur_val: str, prop_val: str, delta_html: str) -> str:
        return f"""
<div style="display:flex;gap:8px;align-items:stretch;flex-direction:column;min-width:130px">
  <div class="kpi kpi-cur">
    <div class="kpi-kind">קיים</div>
    <div class="kpi-val">{cur_val}</div>
    <div class="kpi-lbl">{label}</div>
  </div>
  <div class="kpi kpi-prop">
    <div class="kpi-kind">מוצע</div>
    <div class="kpi-val">{prop_val}</div>
    <div class="kpi-lbl">{label} {delta_html}</div>
  </div>
</div>"""

    summary_cards = ""
    # Allocations
    for key in all_keys:
        lbl   = _ALLOC_LABELS.get(key, key)
        cv    = cur.allocation(key)
        pv    = prop.allocation(key)
        dv    = comparison.delta_allocations.get(key, float("nan"))
        summary_cards += _kpi_pair(lbl, fmt_pct(cv), fmt_pct(pv), _delta_html(dv))

    # Sharpe
    summary_cards += _kpi_pair(
        "שארפ",
        fmt_float(cur.sharpe),
        fmt_float(prop.sharpe),
        _delta_html(comparison.delta_sharpe, suffix=""),
    )
    # Cost
    summary_cards += _kpi_pair(
        "עלות שנתית %",
        fmt_pct(cur.cost, decimals=2),
        fmt_pct(prop.cost, decimals=2),
        _delta_html(comparison.delta_cost, suffix="pp"),
    )

    summary_section = f"""
<div class="section">
  <h2>📋 סיכום מנהלי</h2>
  <div class="kpi-row">{summary_cards}</div>
</div>"""

    # ── 3. Allocation Table ───────────────────────────────────────────────────
    table_rows = ""
    for key in all_keys:
        lbl  = _ALLOC_LABELS.get(key, key)
        cv   = cur.allocation(key)
        pv   = prop.allocation(key)
        dv   = comparison.delta_allocations.get(key, float("nan"))
        table_rows += f"""
<tr>
  <td class="name">{lbl}</td>
  <td>{fmt_pct(cv)}</td>
  <td style="color:#15803d;font-weight:600">{fmt_pct(pv)}</td>
  <td>{_delta_html(dv)}</td>
</tr>"""

    alloc_table = f"""
<table>
  <thead>
    <tr>
      <th class="col-name">רכיב</th>
      <th class="col-cur">קיים</th>
      <th class="col-prop">מוצע</th>
      <th class="col-delta">שינוי</th>
    </tr>
  </thead>
  <tbody>{table_rows}</tbody>
</table>"""

    alloc_section = f"""
<div class="section">
  <h2>📊 השוואת הקצאות</h2>
  {alloc_table}
</div>"""

    # ── 4. Actions Section ────────────────────────────────────────────────────
    action_rows = ""
    _BADGE = {"replace": ("badge-replace", "החלפה"), "add": ("badge-add", "הוספה"), "remove": ("badge-remove", "הסרה")}
    _ICONS = {"replace": "🔄", "add": "✅", "remove": "❌"}

    for act in actions:
        badge_cls, badge_lbl = _BADGE.get(act.action_type, ("", act.action_type))
        icon = _ICONS.get(act.action_type, "•")

        if act.action_type == "replace":
            main_text = f"החלפת <b>{act.original_product}</b> ב-<b>{act.new_product}</b>"
        elif act.action_type == "add":
            main_text = f"הוספת <b>{act.new_product}</b>"
        else:
            main_text = f"הסרת <b>{act.original_product}</b>"

        if act.manager:
            main_text += f" <span style='color:#6b7280;font-size:12px'>({act.manager})</span>"

        impact_html = (
            f"<div class='action-impact'>💡 {act.impact_summary}</div>"
            if act.impact_summary else ""
        )

        action_rows += f"""
<div class="action-row">
  <span class="badge {badge_cls}">{icon} {badge_lbl}</span>
  <div>
    <div class="action-text">{main_text}</div>
    {impact_html}
  </div>
</div>"""

    if not action_rows:
        action_rows = "<p style='color:#6b7280;font-size:13px'>אין פעולות מוצעות.</p>"

    actions_section = f"""
<div class="section">
  <h2>🔧 פעולות מוצעות</h2>
  {action_rows}
</div>"""

    # ── Assemble ──────────────────────────────────────────────────────────────
    body = header + summary_section + alloc_section + actions_section
    footer = f"<div class='footer'>הצעה זו הופקה אוטומטית · {today_str()}</div>"

    return f"<!DOCTYPE html><html dir='rtl' lang='he'><head><meta charset='utf-8'/><title>הצעת תיק השקעות</title>{css}</head><body>{body}{footer}</body></html>"


def build_portfolio_comparison(
    current: "PortfolioSnapshot",
    proposed: "PortfolioSnapshot",
) -> "PortfolioComparison":
    """
    Compare two PortfolioSnapshot objects and return a PortfolioComparison.

    Calculates:
    - delta_allocations : signed pp difference per allocation key
                          (proposed − current); only where both values are known
    - delta_sharpe      : proposed.sharpe − current.sharpe  (nan if either unknown)
    - delta_cost        : proposed.cost  − current.cost     (nan if either unknown)

    Parameters
    ──────────
    current  : PortfolioSnapshot  — the existing / before portfolio
    proposed : PortfolioSnapshot  — the recommended / after portfolio

    Returns
    ───────
    PortfolioComparison with all deltas computed.

    Example
    ───────
    from reporting.report_models import PortfolioSnapshot
    from reporting.report_builder import build_portfolio_comparison

    current = PortfolioSnapshot(
        total_value=1_500_000,
        allocations={"equities": 40.0, "abroad": 30.0, "fx": 25.0, "illiquid": 20.0},
        sharpe=0.82, cost=0.45, managers_count=3, products_count=5,
    )
    proposed = PortfolioSnapshot(
        total_value=1_500_000,
        allocations={"equities": 45.0, "abroad": 27.0, "fx": 25.0, "illiquid": 18.0},
        sharpe=0.90, cost=0.40, managers_count=2, products_count=4,
    )
    comparison = build_portfolio_comparison(current, proposed)
    # comparison.delta_allocations == {"equities": +5.0, "abroad": -3.0, "fx": 0.0, "illiquid": -2.0}
    # comparison.delta_sharpe      == +0.08
    # comparison.delta_cost        == -0.05
    """
    from reporting.report_models import PortfolioComparison

    if not isinstance(current, __import__("reporting.report_models", fromlist=["PortfolioSnapshot"]).PortfolioSnapshot):
        raise TypeError("current חייב להיות PortfolioSnapshot")
    if not isinstance(proposed, __import__("reporting.report_models", fromlist=["PortfolioSnapshot"]).PortfolioSnapshot):
        raise TypeError("proposed חייב להיות PortfolioSnapshot")

    return PortfolioComparison.from_snapshots(current, proposed)
